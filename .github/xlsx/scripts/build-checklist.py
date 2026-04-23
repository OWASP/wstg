import re
import sys
import json
import hashlib
from math import ceil
from copy import copy
from openpyxl import load_workbook
from collections import OrderedDict

checklist_template = '.github/xlsx/assets/checklist-template.xlsx'
workbook = load_workbook(filename=checklist_template)

sheet = workbook.active

def set_version():
    if (len( sys.argv ) > 1):
        return str(sys.argv[1])+" "
    else:
        return ""

def sha256file(file):
    filename = file
    sha256_hash = hashlib.sha256()
    with open(filename,"rb") as f:
        # Read and update hash string value in blocks of 4K
        for byte_block in iter(lambda: f.read(4096),b""):
            sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()

def copy_cell(cellFrom, cellTo):
    cellTo.font = copy(cellFrom.font)
    cellTo.border = copy(cellFrom.border)
    cellTo.fill = copy(cellFrom.fill)
    cellTo.number_format = copy(cellFrom.number_format)
    cellTo.protection = copy(cellFrom.protection)
    cellTo.alignment = copy(cellFrom.alignment)
    cellTo.value = cellFrom.value

def lines_of_text(txt, width):
    lines = txt.split("\n")
    total_lines = 0
    for line in lines:
        total_lines += max(1, ceil(len(line) / float(width)))
    return int(total_lines)

def copy_row(sheet, rowFrom, rowTo):
    for col in range(1, sheet.max_column):
        cellFrom = sheet.cell(row=rowFrom, column=col)
        cellTo = sheet.cell(row=rowTo, column=col)
        copy_cell(cellFrom, cellTo)
    sheet.row_dimensions[rowTo].height = sheet.row_dimensions[rowFrom].height

def set_sheet_title(sheet, version):
    sheet.cell(row=1, column=2).value =  str(sheet.cell(row=1, column=2).value.replace('{version}', version))

def insert_new_header(sheet, title):
    row_template_offset = 3
    col_title_offset = 2
    # sheet.insert_rows(1)
    row = sheet.max_row + 1
    # copy_row(sheet, row_template_offset - 1, row - 1)
    copy_row(sheet, row_template_offset, row)
    sheet.cell(row=row, column=col_title_offset, value=title)

def insert_new_item(sheet, id, name, link, objective):
    row_template_offset = 4
    col_id = 2
    col_name = 3
    col_objective = 4
    col_status = 5

    row = sheet.max_row + 1
    copy_row(sheet, row_template_offset, row)
    sheet.data_validations.dataValidation[0].add(sheet.cell(row=row, column=col_status))

    # Adjust height
    max_lines = max(
        lines_of_text(name, 40),
        lines_of_text(objective, 60))

    # Set values
    sheet.cell(row=row, column=col_id, value=id)
    sheet.cell(row=row, column=col_name, value='=HYPERLINK("{}", "{}")'.format(link, name))
    sheet.cell(row=row, column=col_objective, value=objective)

    base_height = sheet.row_dimensions[row_template_offset].height
    sheet.row_dimensions[row].height = max_lines * base_height

def insert_empty(sheet):
    row_template_offset = 2
    row = sheet.max_row + 1
    copy_row(sheet, row_template_offset, row)

# Load Checklist
checklist_json_path = 'checklists/checklist.json'
checklist_readme_path = 'checklists/README.md'
checklist_output_file = 'checklists/checklist.xlsx'
checklist = json.load(open(checklist_json_path), object_pairs_hook=OrderedDict)

# Set Version
version = set_version()

# Set Sheet Title
set_sheet_title(sheet, version)

for categoryKey in checklist['categories']:
    print("Generate category: " + categoryKey)
    category = checklist['categories'][categoryKey]

    insert_new_header(sheet, categoryKey)

    for item in category['tests']:
        separator = "- "
        emptyObjective = item['objectives'][0] == ""
        objective = "N/A" if emptyObjective else separator + ("\n"+separator).join(item['objectives'])
        insert_new_item(sheet,
            id=item['id'],
            name=item['name'],
            link=item['reference'],
            objective=objective)
    insert_empty(sheet)

# Replicates conditional formatting
cf = list(sheet.conditional_formatting._cf_rules.keys())[0]
rules = sheet.conditional_formatting._cf_rules[cf]
cf_range = "B4:F"+str(sheet.max_row)
for rule in rules:
    new_rule = copy(rule)
    sheet.conditional_formatting.add(cf_range, copy(rule))

# Remove template rows
# sheet.delete_rows(3)
copy_row(sheet, 2, 3)
copy_row(sheet, 2, 4)
sheet.row_dimensions[3].height = 5
sheet.row_dimensions[4].height = 5

workbook.save(filename=checklist_output_file)

# Check sha256 and modifies README
file_sha256 = sha256file(checklist_output_file)
readme_content = ''
with open (checklist_readme_path, 'r' ) as f:
    content = f.read()
    readme_content = re.sub('SHA-256: (\w+)', 'SHA-256: '+file_sha256, content, flags = re.M)

with open (checklist_readme_path, 'w') as f:
    f.write(readme_content)